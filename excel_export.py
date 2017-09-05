from appJar import gui
import openpyxl
from openpyxl import worksheet as ws
import re
import os
import datetime
import pandas as pd
import numpy as np
buttons = [
    'Go!',
    'Cancel'
]
# TODO create regex to check file names per report
rec_activity_Regex = re.compile(
    r'Recognition.Activity+'
)
hris_report_Regex = re.compile(
    r'Export.Members+'
)
app = gui(title="CS Report Utility")
file_name = "*CS_Compiled_Reports_" + str(datetime.date.today()) + ".xlsx"


def copy_worksheet_values(source_worksheet: ws.Worksheet, target_worksheet: ws.Worksheet):
    """
    Copies cell values from the source worksheet to the target worksheet


    Keyword arguments:
    source_worksheet -- the source (starting) worksheet to copy from
    target_worksheet -- the target worksheet to copy data to
    """
    rows = []
    for row in source_worksheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    for source_rows in rows:
        target_worksheet.append(source_rows)


def launch(button):
    if button == "Cancel":
        quit()
    elif not app.getEntry(name="find_dir"):
        msg = "You need to pick a directory to start the app."
        app.warningBox(title="ERROR: Missing Information", message=msg)
    else:
        direc = app.getEntry(name="find_dir")
        # check the chosen directory for # of files
        entry = os.scandir(direc)
        file_cnt = 0
        filelist = []
        for i in entry:
            if not i.name.startswith('.') and i.is_file():
                # if file is csv, warn user the utility doesn't accept CSV
                # TODO - add handling that CSV files are converted instead of errored out
                if os.path.splitext(i)[1] == '.csv':
                    msg = "Utility can only take Excel files. This file is a .csv: "
                    app.infoBox(title="Critical Error", message=msg + i.name)
                    quit()
                file_cnt += 1
                filelist.append(i.name)
        # create workbook
        wb = openpyxl.Workbook()
        # count each loop for file naming
        counter = 1
        # copy/paste from other xlsx into new file
        for excel in filelist:
            # create a new sheet per file (uses default naming)
            wb_sheet = wb.create_sheet(title="HG Report " + str(counter))
            # load saved off workbook based on selected directory
            load_wb = openpyxl.load_workbook(os.path.join(direc, excel))
            # pull the active sheet
            copy_worksheet_values(source_worksheet=load_wb.active, target_worksheet=wb_sheet)
            counter += 1
        wb.save(filename=(os.path.join(direc, file_name)))
        finish_up(file_cnt=file_cnt)
        quit()


def main():
    msg = "Runs calculations over multiple spreadsheets. Select where all the source spreadsheets are stored.\n"
    msg2 = "A spreadsheet will be created in that folder combining this data."
    app.addLabel(title="header", text=msg + msg2)
    app.addHorizontalSeparator()
    app.addDirectoryEntry(title="find_dir", )
    app.addButtons(names=buttons, funcs=launch)
    app.go()


def calculate():
    quit()


def finish_up(file_cnt):
    comp_msg = "Your workbook is complete.\n File named: " + file_name + '\n'
    total_msg = "Spreadsheets combined: "
    app.infoBox(title="Workbook Complete!", message=comp_msg + total_msg + str(file_cnt))


if __name__ == '__main__':
    main()
